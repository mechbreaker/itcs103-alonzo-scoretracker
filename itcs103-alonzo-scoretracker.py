import tkinter as tk
from tkinter import *
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os
from tkinter import ttk
from tkinter import messagebox

wb = Workbook()
ws = wb.active
ws.title = "Student Score Tracker"
wb = load_workbook(r"python/itcs103-alonzo-scoretracker/score.xlsx")
ws = wb["Student Score Tracker"]
cell = ws.cell(row=1, column=1)
if cell.value is None:
    ws.append(['name', 'grade', 'score'])
    wb.save(r"python/itcs103-alonzo-scoretracker/score.xlsx")



window = tk.Tk()
window.geometry("350x400")
window.title("excel moment")
window.eval('tk::PlaceWindow . center')
rowcount=0


#table
table = ttk.Treeview(window)
table['columns'] = ('name', 'grade', 'score')

table.column('#0', width=0, stretch=tk.NO)
table.column('name', anchor=tk.CENTER, width=150)
table.column('grade', anchor=tk.CENTER, width=75)
table.column('score', anchor=tk.CENTER, width=75)
#############################################################
table.heading('#0', text='', anchor=tk.CENTER)
table.heading('name', text='name', anchor=tk.CENTER)
table.heading('grade', text='grade', anchor=tk.CENTER)
table.heading('score', text='score', anchor=tk.CENTER)

data = [
]

table.tag_configure('oddrow', background='#E8E8E8')
table.tag_configure('evenrow', background='#FFFFFF')

for i in range(len(data)):
    if i % 2 == 0:
        table.insert(parent='', index=i, values=data[i], tags=('evenrow',))
    else:
        table.insert(parent='', index=i, values=data[i], tags=('oddrow',))



name=tk.Entry(window)
grade=tk.Entry(window)
name.grid(row=2, column=0, sticky="nsew", padx=5,pady=5)
grade.grid(row=2, column=1, sticky="nsew", padx=5,pady=5)


def validate_inputs():
    studentname = name.get()
    score_text = grade.get()

    if not studentname or not score_text:
        messagebox.showerror("Error", "All fields are required!")
        return False

    try:
        float(score_text)
    except ValueError:
        messagebox.showerror("Error", "Score must be a number!")
        return False

    return True

def format_excel(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "Average":
            for cell in row:
                cell.font = Font(bold=True)
            break
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2


def save_to_excel():
    if not validate_inputs():
        return

    studentname = name.get()
    score = float(grade.get())
    if score >100:
        messagebox.showinfo("Error", "maximum grade input is 100")
        grade.delete(0, tk.END)
    elif score < 0:
        messagebox.showinfo("Error", "invalid number")
        grade.delete(0, tk.END)
    else:
        wb = load_workbook(r"python/itcs103-alonzo-scoretracker/score.xlsx")
        ws = wb["Student Score Tracker"]

        for row in ws.iter_rows(min_row=2):
            if row[0].value == "Average":
                ws.delete_rows(row[0].row)
                break

        remarks = "Pass" if score >= 60 else "Fail"
        ws.append([studentname, score, remarks])

        scores = [cell.value for cell in ws["B"] if isinstance(cell.value, (int, float))]
        if scores:
            average = sum(scores) / len(scores)
            ws.append(["Average", average, ""])

        format_excel(ws)
        wb.save(r"python/itcs103-alonzo-scoretracker/score.xlsx")

        messagebox.showinfo("Success", "Data saved to Excel!")
        name.delete(0, tk.END)
        grade.delete(0, tk.END)
        show_data()

data_window = None
datacount=[]

def show_data():
    global data_window

    wb = load_workbook(r"python/itcs103-alonzo-scoretracker/score.xlsx")
    ws = wb["Student Score Tracker"]

    # reloads table if saved data exists
    if table.get_children() != ():
        for i in table.get_children():
            table.delete(i)
            print('replaced')

    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header row
        table.insert("", tk.END, values=row)
        
    
    for i in table.get_children():
        global rowcount
        rowcount+=1
        datacount.append(f"A{rowcount}")
    
    print(len(datacount))
 
def clear_data():
    print()
    a= table.get_children()
    print(a)
show_data()
table.grid(row=0, column=0, sticky="nsew" ,columnspan=2, padx=5,pady=5)


#buttons
lblname=Label(window, text="name").grid(row=1, column=0, sticky="nsew", padx=5,pady=5)

lblgrade=Label(window, text="grade").grid(row=1, column=1, sticky="nsew", padx=5,pady=5)


submit = tk.Button(window, text="submit", command=save_to_excel).grid(row=3, column=0, sticky="nsew", padx=5,pady=5)
edit = tk.Button(window, text="edit (under maintenance)", command=clear_data).grid(row=3, column=1, sticky="nsew", padx=5,pady=5)


##functions

# Create a Combobox widget

combobox = ttk.Combobox(window, values=datacount, state="readonly")
combobox.grid(row=4, column=0, sticky="nsew", padx=5,pady=5)

# Set a default value
combobox.set("Select an option")
window.mainloop()